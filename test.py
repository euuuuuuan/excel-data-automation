import pandas as pd
import glob
from openpyxl import load_workbook


def update_product_list_with_format():
    """
    상품리스트 엑셀 파일의 양식을 유지하며
    '전체숨김여부'를 'off'로 수정하고 새 파일로 저장하는 함수
    """

    # 파일명 정의
    product_list_file = '더미_상품리스트.xlsx'
    monthly_files_pattern = '더미_*월_정산내역.xlsx'
    output_file = '더미_상품리스트_업데이트.xlsx'

    try:
        # 원본 엑셀 파일을 양식 그대로 불러옵니다.
        wb = load_workbook(product_list_file)
        ws = wb['Sheet1']  # 'Sheet1' 시트 선택
    except FileNotFoundError:
        print(f"오류: 상품리스트 파일 '{product_list_file}'을 찾을 수 없습니다. 경로와 파일명을 확인해주세요.")
        return
    except KeyError:
        print("오류: 'Sheet1' 시트를 찾을 수 없습니다. 시트 이름을 확인해 주세요.")
        return
    except Exception as e:
        print(f"상품리스트 파일 로드 중 오류 발생: {e}")
        return

    print("--- 엑셀 파일 비교 및 수정 시작 ---")

    # 월별 정산내역 파일들을 순회하며 상품코드 추출
    monthly_files = glob.glob(monthly_files_pattern)
    if not monthly_files:
        print("오류: 정산내역 파일을 찾을 수 없습니다.")
        return

    found_product_codes = set()
    for file_path in monthly_files:
        print(f"'{file_path}' 파일 처리 중...")
        try:
            xls = pd.ExcelFile(file_path)
            for sheet_name in xls.sheet_names:
                # 월별 파일의 헤더 위치는 1행이므로 header=0으로 유지
                df_monthly = pd.read_excel(file_path, sheet_name=sheet_name, header=0)

                cols = {str(col).lower(): col for col in df_monthly.columns}
                if '상품코드' in cols:
                    product_code_column_name = cols['상품코드']
                    codes = df_monthly[product_code_column_name].dropna().astype(str).tolist()
                    found_product_codes.update(codes)

        except Exception as e:
            print(f"'{file_path}' 파일 처리 중 오류 발생: {e}")
            continue

    print(f"\n총 {len(found_product_codes)}개의 고유 상품코드를 정산내역 파일에서 찾았습니다.")

    # openpyxl로 상품리스트 파일의 '상품코드'와 '전체숨김여부' 열의 위치를 찾습니다.
    # 상품리스트 파일의 헤더는 2행이므로 header_row를 2로 설정합니다.
    header_row = 2
    product_code_col_idx = None
    status_col_idx = None

    # 헤더 행에서 열 이름 찾기
    for cell in ws[header_row]:
        if cell.value == '상품코드':
            product_code_col_idx = cell.column
        if cell.value == '전체숨김여부':
            status_col_idx = cell.column

    if not product_code_col_idx or not status_col_idx:
        print("오류: '상품코드' 또는 '전체숨김여부' 열을 엑셀 파일에서 찾을 수 없습니다.")
        return

    # 상품리스트 파일의 '전체숨김여부' 값을 수정합니다.
    # 데이터는 3행부터 시작
    for row in range(header_row + 1, ws.max_row + 1):
        product_code_cell = ws.cell(row=row, column=product_code_col_idx)
        status_cell = ws.cell(row=row, column=status_col_idx)

        # 상품코드가 found_product_codes 집합에 있으면 'off'로 수정
        if product_code_cell.value is not None and str(product_code_cell.value) in found_product_codes:
            status_cell.value = 'off'

    # 수정된 내용을 새로운 파일로 저장합니다.
    try:
        wb.save(output_file)
        print(f"\n✔️ '{output_file}' 파일이 성공적으로 생성되었습니다.")
    except Exception as e:
        print(f"\n오류: 수정된 파일 저장 중 오류 발생: {e}")

# 함수 실행
update_product_list_with_format()