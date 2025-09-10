import glob
from openpyxl import load_workbook
import uuid
import random


def create_dummy_files():
    """
    원본 엑셀 파일의 구조와 서식을 유지하면서 민감 정보를 제거하고,
    상품리스트와 정산내역 파일 간의 상품 코드를 완벽하게 연동시켜 더미 파일을 생성하는 함수
    """

    print("--- 더미 파일 생성 시작 ---")

    # 1. 상품리스트 파일 더미 생성
    product_list_file = '이디엠에듀케이션_상품리스트(250826).xlsx'
    dummy_product_list_file = '더미_상품리스트.xlsx'

    # 상품리스트 파일에 사용될 더미 상품 코드를 모두 미리 생성
    dummy_product_codes = []

    try:
        wb = load_workbook(product_list_file)
        ws = wb['Sheet1']
    except Exception as e:
        print(f"오류: '{product_list_file}' 파일 로드 실패. {e}")
        return

    header_row_pl = 2
    product_code_col_pl = None
    product_name_col_pl = None
    status_col_pl = None
    for cell in ws[header_row_pl]:
        if cell.value == '상품코드':
            product_code_col_pl = cell.column
        if cell.value == '상품명':
            product_name_col_pl = cell.column
        if cell.value == '전체숨김여부':
            status_col_pl = cell.column

    if not product_code_col_pl or not product_name_col_pl or not status_col_pl:
        print("오류: '상품코드', '상품명' 또는 '전체숨김여부' 열을 상품리스트 파일에서 찾을 수 없습니다.")
        return

    for row_idx in range(header_row_pl + 1, ws.max_row + 1):
        # 상품리스트용 더미 상품 코드 생성 및 리스트에 추가
        dummy_code = str(uuid.uuid4())[:8]
        ws.cell(row=row_idx, column=product_code_col_pl, value=dummy_code)
        ws.cell(row=row_idx, column=product_name_col_pl, value=f'더미상품_{row_idx - header_row_pl}')
        ws.cell(row=row_idx, column=status_col_pl, value='on')
        dummy_product_codes.append(dummy_code)

    wb.save(dummy_product_list_file)
    print(f"✔️ '{dummy_product_list_file}' 파일이 성공적으로 생성되었습니다.")

    # 2. 월별 정산내역 파일 더미 생성
    monthly_files_pattern = '이디엠에듀케이션_*월_정산내역.xlsx'
    monthly_files = glob.glob(monthly_files_pattern)

    sensitive_cols = ['고객코드', '고객명', '로그인ID', '주문번호', '상품코드', '상품명', '단위', '수량', '단가', '공급가', '부가세', '합계', '회원명',
                      '매출일자', '매출번호', '수령자명']

    # 월별 정산내역에 사용될 상품 코드 리스트
    # 상품리스트에서 생성된 코드를 복사하고 무작위로 섞음
    monthly_sales_codes = list(dummy_product_codes)
    random.shuffle(monthly_sales_codes)

    # 순환 가능한 이터레이터로 만들어 코드 재사용
    from itertools import cycle
    code_cycler = cycle(monthly_sales_codes)

    for file_path in monthly_files:
        dummy_file_path = f"더미_{file_path}"
        try:
            wb = load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                print(f"'{file_path}'의 시트 '{sheet_name}' 처리 중...")

                header_row = -1
                for row_idx in range(1, 10):
                    cell_values = [str(cell.value) if cell.value is not None else '' for cell in ws[row_idx]]
                    if any(col in cell_values for col in sensitive_cols):
                        header_row = row_idx
                        break

                if header_row == -1:
                    print(f"경고: 시트 '{sheet_name}'에서 민감 정보 헤더를 찾을 수 없습니다. 건너뜁니다.")
                    continue

                col_indices = {col: None for col in sensitive_cols}
                for cell in ws[header_row]:
                    if cell.value in col_indices:
                        col_indices[cell.value] = cell.column

                for row_idx in range(header_row + 1, ws.max_row + 1):
                    for col_name, col_idx in col_indices.items():
                        if col_idx:
                            if col_name == '상품명':
                                ws.cell(row=row_idx, column=col_idx, value=f'더미상품_{row_idx - header_row}')
                            elif col_name == '상품코드':
                                # 상품리스트의 코드를 순환하며 할당
                                ws.cell(row=row_idx, column=col_idx, value=next(code_cycler))
                            else:
                                ws.cell(row=row_idx, column=col_idx, value='')

            wb.save(dummy_file_path)
            print(f"✔️ '{dummy_file_path}' 파일이 성공적으로 생성되었습니다.")
        except Exception as e:
            print(f"오류: '{file_path}' 파일 처리 실패. {e}")


# 함수 실행
create_dummy_files()